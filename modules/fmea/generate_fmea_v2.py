"""
Drone Propulsion System FMEA Generator
Produces: propulsion_fmea.xlsx, propulsion_fmea.csv, and markdown reports
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from datetime import date, timedelta
import random

random.seed(7)
from pathlib import Path

ROOT_DIR = Path(__file__).resolve().parents[2]

DATA_DIR = ROOT_DIR / "data"
REPORTS_DIR = ROOT_DIR / "reports"

DATA_DIR.mkdir(exist_ok=True)
REPORTS_DIR.mkdir(exist_ok=True)

# ─── FMEA DATA ───────────────────────────────────────────────────────────────

today = date.today()

def d(offset): return (today + timedelta(days=offset)).isoformat()

FMEA_ROWS = [
    # ── MOTOR ──────────────────────────────────────────────────────────────
    (1,"Propulsion","Motor","Convert electrical energy to rotational thrust",
     "Bearing seizure","Complete loss of motor rotation → vehicle crash",
     9,"Insufficient lubrication / contamination during assembly",5,
     "Assembly SOP with torque spec","End-of-line vibration test",5,225,
     "Add bearing inspection step at motor install; specify grease type in BOM",
     "QE-Torres",d(30),9,3,3,81,"Open"),
    (2,"Propulsion","Motor","Convert electrical energy to rotational thrust",
     "Rotor imbalance","Excessive vibration → structural fatigue of mount bracket",
     7,"Damaged or contaminated rotor blades during handling",4,
     "Incoming visual inspection","End-of-line vibration signature check",4,112,
     "Implement dynamic balance check at motor receiving inspection",
     "QE-Patel",d(21),7,2,2,28,"In Progress"),
    (3,"Propulsion","Motor","Convert electrical energy to rotational thrust",
     "Winding insulation failure","Electrical short → ESC shutdown → propulsion loss",
     9,"Thermal cycling stress; manufacturing defect",3,
     "Supplier IPC-A-610 cert required","Hi-pot dielectric test at EOL",3,81,
     "Require 100% hi-pot test at supplier; add thermal shock screen",
     "QE-Nguyen",d(14),9,2,2,36,"In Progress"),
    (4,"Propulsion","Motor","Convert electrical energy to rotational thrust",
     "Motor overheating","Demagnetization → power loss; potential fire risk",
     8,"Incorrect motor specified for duty cycle; blocked airflow",4,
     "Thermal simulation in design review","Thermal imaging at end-of-line",4,128,
     "Add thermal derating curve to motor spec; test at max load",
     "QE-Torres",d(45),8,2,3,48,"Open"),
    (5,"Propulsion","Motor","Convert electrical energy to rotational thrust",
     "Incorrect motor installed","Wrong thrust profile → flight control failure",
     8,"Part number confusion on similar-looking variants",3,
     "Part number label on motor","Visual check at installation",5,120,
     "Implement barcode scan verification at motor installation station",
     "QE-Brooks",d(10),8,1,2,16,"Complete"),

    # ── ESC ────────────────────────────────────────────────────────────────
    (6,"Propulsion","Electronic Speed Controller","Regulate motor power via PWM",
     "Firmware mismatch","Motor runs at wrong speed profile → unstable flight",
     8,"Firmware version not validated against motor/FC combination",5,
     "Firmware release notes reviewed at build","EOL motor spin-up test",4,160,
     "Lock firmware version in build traveler; add version check to EOL script",
     "QE-Osei",d(14),8,2,2,32,"In Progress"),
    (7,"Propulsion","Electronic Speed Controller","Regulate motor power via PWM",
     "Thermal shutdown","Mid-flight power cutoff → crash",
     9,"Inadequate heatsinking; ambient temperature exceeded spec",4,
     "Thermal analysis at design stage","Thermal imaging at EOL",4,144,
     "Add 10°C thermal margin to ESC mounting design; validate in thermal chamber",
     "QE-Torres",d(30),9,2,3,54,"In Progress"),
    (8,"Propulsion","Electronic Speed Controller","Regulate motor power via PWM",
     "Signal loss","Motor stops receiving throttle signal → vehicle drop",
     10,"Wiring harness damage; connector vibration loosening",4,
     "Wiring harness routing guidelines","EOL signal continuity test",3,120,
     "Add locking connector on ESC signal line; vibration test at system level",
     "QE-Patel",d(21),10,2,2,40,"In Progress"),
    (9,"Propulsion","Electronic Speed Controller","Regulate motor power via PWM",
     "Solder joint failure","Intermittent / complete power loss",
     7,"Poor wetting; thermal fatigue from power cycling",4,
     "Supplier IPC-A-610 Class 3 required","X-ray inspection sample plan",5,140,
     "Increase x-ray sample rate to 20%; add thermal cycle screen",
     "QE-Nguyen",d(21),7,3,3,63,"Open"),
    (10,"Propulsion","Electronic Speed Controller","Regulate motor power via PWM",
     "Incorrect calibration","Motor timing mismatch → efficiency loss / instability",
     6,"Calibration step skipped or documented incorrectly",3,
     "Calibration procedure in work instruction","EOL calibration verification",3,54,
     "Add mandatory calibration step with pass/fail criteria to build traveler",
     "QE-Brooks",d(7),6,2,2,24,"Complete"),

    # ── PROPELLER ──────────────────────────────────────────────────────────
    (11,"Propulsion","Propeller","Generate thrust via aerodynamic lift",
     "Crack","Blade separation → catastrophic imbalance → crash",
     10,"Impact damage; material defect; fatigue from over-torque",3,
     "Incoming visual inspection SOP","EOL vibration signature test",4,120,
     "Add dye-penetrant inspection to receiving; torque limiter on installation tool",
     "QE-Torres",d(14),10,2,2,40,"In Progress"),
    (12,"Propulsion","Propeller","Generate thrust via aerodynamic lift",
     "Delamination","Loss of aerodynamic profile → thrust asymmetry",
     8,"CFRP layup defect; moisture ingress during storage",4,
     "Supplier process audit annually","Ultrasonic C-scan sample inspection",5,160,
     "Add C-scan to receiving AQL plan for CFRP propellers; improve storage spec",
     "QE-Osei",d(30),8,2,3,48,"Open"),
    (13,"Propulsion","Propeller","Generate thrust via aerodynamic lift",
     "Incorrect pitch installed","Wrong performance curve → under/over thrust",
     7,"Wrong part pulled from inventory; no pitch marking check",4,
     "Part labeling standard","Visual check at assembly",5,140,
     "Add pitch laser-etch marking; implement scan-to-verify at assembly",
     "QE-Patel",d(10),7,2,2,28,"Complete"),
    (14,"Propulsion","Propeller","Generate thrust via aerodynamic lift",
     "Loose attachment","Propeller detaches in flight → crash",
     10,"Under-torque at installation; incorrect fastener",2,
     "Torque spec in work instruction","Torque audit at EOL",2,40,
     "Use torque-indicating fasteners; add torque verification to build record",
     "QE-Brooks",d(7),10,1,1,10,"Complete"),
    (15,"Propulsion","Propeller","Generate thrust via aerodynamic lift",
     "Surface damage","Drag increase → reduced endurance",
     4,"Handling damage post-inspection",3,
     "Handling guidelines in SOP","Visual inspection before installation",4,48,
     "Add foam-lined storage trays; inspect immediately before installation",
     "QE-Nguyen",d(14),4,2,3,24,"Open"),

    # ── MOTOR MOUNT BRACKET ────────────────────────────────────────────────
    (16,"Propulsion","Motor Mount Bracket","Rigidly locate motor to airframe",
     "Hole position out of tolerance","Motor misalignment → vibration → fatigue",
     7,"Fixture wear; CNC program error",5,
     "First article inspection (FAI) required","CMM inspection per AQL",5,175,
     "Tighten AQL from 1.0 to 0.4 for hole position; add SPC on CNC fixture",
     "QE-Torres",d(21),7,3,3,63,"In Progress"),
    (17,"Propulsion","Motor Mount Bracket","Rigidly locate motor to airframe",
     "Crack at mounting feature","Structural failure → motor detachment → crash",
     9,"Stress concentration; incorrect alloy; over-torque",3,
     "FEA review in design; material cert required","100% visual + dye-pen on critical features",3,81,
     "Add stress-relief radius to design; specify 7075-T6 only in BOM",
     "QE-Patel",d(14),9,2,2,36,"In Progress"),
    (18,"Propulsion","Motor Mount Bracket","Rigidly locate motor to airframe",
     "Flatness failure","Poor motor contact → vibration transmission",
     6,"Machining stress relief insufficient",4,
     "Flatness spec on drawing","CMM flatness measurement",3,72,
     "Add stress-relief annealing step to manufacturing router",
     "QE-Nguyen",d(21),6,2,2,24,"Complete"),
    (19,"Propulsion","Motor Mount Bracket","Rigidly locate motor to airframe",
     "Fastener interface wear","Loose motor over time → vibration increase",
     5,"Insufficient thread engagement; wrong fastener grade",4,
     "Fastener spec in BOM","Torque-off test at sample rate",4,80,
     "Specify thread insert (Helicoil) for high-cycle fastener interfaces",
     "QE-Brooks",d(30),5,2,3,30,"Open"),
    (20,"Propulsion","Motor Mount Bracket","Rigidly locate motor to airframe",
     "Incorrect material lot","Wrong alloy → reduced strength margin",
     8,"Material certs not verified at receiving",3,
     "Supplier material cert required","PMI (positive material identification) sample",5,120,
     "Add 100% PMI scan at receiving inspection for motor mount brackets",
     "QE-Osei",d(14),8,2,2,32,"In Progress"),

    # ── WIRING HARNESS ─────────────────────────────────────────────────────
    (21,"Propulsion","Wiring Harness","Transmit power and signal to motor/ESC",
     "Connector not fully seated","Intermittent power/signal → erratic motor behavior",
     8,"Assembly error; connector design requires high insertion force",5,
     "Assembly SOP with seating verification","EOL continuity test",4,160,
     "Add positive-lock connectors; add seating force verification to work instruction",
     "QE-Torres",d(14),8,3,2,48,"In Progress"),
    (22,"Propulsion","Wiring Harness","Transmit power and signal to motor/ESC",
     "Wire chafing","Short circuit → fire risk or power loss",
     9,"Inadequate routing protection in high-vibration zones",4,
     "Routing guidelines in design SOP","Visual inspection at final assembly",5,180,
     "Add grommets/abrasion sleeves at all frame contact points; add to inspection checklist",
     "QE-Patel",d(21),9,2,3,54,"Open"),
    (23,"Propulsion","Wiring Harness","Transmit power and signal to motor/ESC",
     "Incorrect pinout","Motor runs in reverse or not at all",
     7,"Wiring error; missing polarization feature",3,
     "Pinout drawing on work instruction","Functional motor spin test at EOL",3,63,
     "Add polarized connector shrouds; implement spin-direction check in EOL software",
     "QE-Nguyen",d(10),7,1,2,14,"Complete"),
    (24,"Propulsion","Wiring Harness","Transmit power and signal to motor/ESC",
     "Insulation damage","Potential short; degraded signal integrity",
     7,"Handling damage; sharp edge contact",4,
     "Handling SOP","Visual inspection at installation",4,112,
     "Add insulation damage to receiving inspection AQL; improve packaging spec",
     "QE-Brooks",d(21),7,2,3,42,"Open"),
    (25,"Propulsion","Wiring Harness","Transmit power and signal to motor/ESC",
     "Intermittent connection","Random power loss → flight instability",
     8,"Vibration-induced fretting; poor crimp quality",5,
     "Crimp quality spec in supplier requirements","EOL vibration + continuity test",4,160,
     "Add crimp pull-test to receiving inspection; specify vibration-resistant crimp termination",
     "QE-Osei",d(30),8,3,2,48,"In Progress"),

    # ── BATTERY CONNECTOR ──────────────────────────────────────────────────
    (26,"Propulsion","Battery Connector","Connect battery to power distribution",
     "Voltage drop","Reduced power delivery → ESC brownout",
     7,"High contact resistance; oxidation; undersized contact",4,
     "Connector rated spec in BOM","Voltage drop test at EOL",3,84,
     "Add contact resistance measurement to receiving AQL",
     "QE-Torres",d(14),7,2,2,28,"In Progress"),
    (27,"Propulsion","Battery Connector","Connect battery to power distribution",
     "Loose contact","Intermittent power → erratic flight behavior",
     8,"Worn contact spring; incorrect mating cycle count exceeded",4,
     "Mating cycle rating in connector spec","Manual wiggle test + EOL continuity",5,160,
     "Add mating cycle tracker to battery swap log; replace connectors at spec limit",
     "QE-Patel",d(21),8,3,3,72,"Open"),
    (28,"Propulsion","Battery Connector","Connect battery to power distribution",
     "Thermal damage","Connector melts → fire or power loss",
     9,"Current draw exceeds connector rating; poor heat dissipation",3,
     "Connector ampacity spec in BOM","Thermal imaging at EOL under load",3,81,
     "Uprate connector to next ampacity class; add thermal fuse in line",
     "QE-Nguyen",d(14),9,2,2,36,"In Progress"),
    (29,"Propulsion","Battery Connector","Connect battery to power distribution",
     "Reverse polarity risk","Battery connected backwards → immediate ESC/motor damage",
     10,"Symmetric connector allows incorrect mating",2,
     "Polarity labeling on connector and battery","Visual polarity check in work instruction",4,80,
     "Replace with asymmetric keyed connector to physically prevent reverse polarity",
     "QE-Brooks",d(7),10,1,1,10,"Complete"),
    (30,"Propulsion","Battery Connector","Connect battery to power distribution",
     "Contamination at contact surface","Increased resistance → heating → degraded performance",
     6,"Exposure to oils, dust, moisture in assembly environment",4,
     "Cleanliness spec in assembly area","Visual inspection at assembly",4,96,
     "Add connector cap/plug until mating; add cleaning step to assembly SOP",
     "QE-Osei",d(21),6,2,3,36,"Open"),

    # ── FASTENERS ──────────────────────────────────────────────────────────
    (31,"Propulsion","Fasteners","Mechanically secure propulsion components",
     "Under-torque on critical joint","Joint loosens under vibration → component shift",
     7,"Torque wrench not calibrated; SOP not followed",4,
     "Calibrated torque tools; torque spec in SOP","Torque audit at final assembly",3,84,
     "Add digital torque tool with data logging; flag out-of-spec torques in real time",
     "QE-Torres",d(14),7,2,2,28,"In Progress"),
    (32,"Propulsion","Fasteners","Mechanically secure propulsion components",
     "Wrong fastener grade installed","Insufficient clamp load → joint failure under load",
     7,"Mixed inventory; no grade marking on small fasteners",4,
     "Part number on fastener bags","Visual check at assembly",4,112,
     "Implement kitting by job traveler; add grade stamp requirement to fastener spec",
     "QE-Patel",d(21),7,2,3,42,"Open"),

    # ── THERMAL INTERFACE MATERIAL ─────────────────────────────────────────
    (33,"Propulsion","Thermal Interface Material","Conduct heat from ESC to heatsink",
     "Incorrect thickness applied","Thermal resistance above design spec → overheating",
     6,"Manual application variability; no thickness gauge used",5,
     "TIM application procedure in work instruction","Thermal imaging at EOL",4,120,
     "Add shim template for TIM application; add thickness check with feeler gauge",
     "QE-Nguyen",d(14),6,3,3,54,"Open"),
    (34,"Propulsion","Thermal Interface Material","Conduct heat from ESC to heatsink",
     "Dry-out / delamination over thermal cycles","Thermal resistance increases with age",
     5,"Low-grade TIM material; exceeding max operating temp",3,
     "TIM material spec requires thermal conductivity rating","Thermal test at qualification",4,60,
     "Upgrade to phase-change TIM; add thermal cycle qualification test",
     "QE-Brooks",d(30),5,2,3,30,"Open"),

    # ── BATTERY CONNECTOR (additional) ─────────────────────────────────────
    (35,"Propulsion","Wiring Harness","Transmit power and signal to motor/ESC",
     "Over-length wire routing","Wire caught in rotating propeller → crash",
     10,"No wire length specification; inconsistent assembly",3,
     "Wire routing diagram in work instruction","Visual pre-flight inspection checklist",3,90,
     "Add wire length to BOM; add wire-to-prop clearance check to EOL inspection",
     "QE-Osei",d(7),10,2,1,20,"In Progress"),
]

COLUMNS = [
    "fmea_id","subsystem","component","function","failure_mode","failure_effect",
    "severity","potential_cause","occurrence","current_prevention_control",
    "current_detection_control","detection","rpn","recommended_action",
    "action_owner","target_completion_date","revised_severity","revised_occurrence",
    "revised_detection","revised_rpn","status"
]

df = pd.DataFrame(FMEA_ROWS, columns=COLUMNS)
df.to_csv(DATA_DIR / "propulsion_fmea.csv", index=False)
print(f"✓ propulsion_fmea.csv  ({len(df)} rows)")

# ─── EXCEL FILE ──────────────────────────────────────────────────────────────

wb = Workbook()

# ── Sheet 1: FMEA Main ──────────────────────────────────────────────────────
ws = wb.active
ws.title = "FMEA"

HEADER_FILL  = PatternFill("solid", fgColor="1F2D3D")
SUBHDR_FILL  = PatternFill("solid", fgColor="2D4159")
ALT_FILL     = PatternFill("solid", fgColor="0D1117")
ALT_FILL2    = PatternFill("solid", fgColor="161B22")
RED_FILL     = PatternFill("solid", fgColor="F85149")
AMBER_FILL   = PatternFill("solid", fgColor="D29922")
GREEN_FILL   = PatternFill("solid", fgColor="3FB950")
BLUE_FILL    = PatternFill("solid", fgColor="388BFD")

WHITE  = Font(color="F0F6FC", bold=True, name="Arial", size=9)
NORMAL = Font(color="C9D1D9", name="Arial", size=9)
BOLD   = Font(color="F0F6FC", bold=True, name="Arial", size=9)

thin = Side(style="thin", color="21262D")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Title row
ws.merge_cells("A1:U1")
ws["A1"] = "DRONE PROPULSION SYSTEM — FAILURE MODE AND EFFECTS ANALYSIS (FMEA)"
ws["A1"].font = Font(color="58A6FF", bold=True, name="Arial", size=13)
ws["A1"].fill = PatternFill("solid", fgColor="0D1117")
ws["A1"].alignment = CENTER
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:U2")
ws["A2"] = f"Part Family: Drone Propulsion System   |   Revision: A   |   Date: {today.isoformat()}   |   Owner: Quality Engineering"
ws["A2"].font = Font(color="8B949E", name="Arial", size=9)
ws["A2"].fill = PatternFill("solid", fgColor="0D1117")
ws["A2"].alignment = CENTER
ws.row_dimensions[2].height = 18

# Section headers (row 3)
SECTION_HEADERS = [
    ("A3:A3","ID"), ("B3:D3","COMPONENT INFORMATION"), ("E3:F3","FAILURE ANALYSIS"),
    ("G3:I3","RISK ASSESSMENT (BEFORE)"), ("J3:L3","CURRENT CONTROLS"),
    ("M3:M3","RPN"), ("N3:O3","RECOMMENDED ACTION"),
    ("P3:P3","TARGET DATE"), ("Q3:T3","REVISED RISK (AFTER)"), ("U3:U3","STATUS"),
]
for span, label in SECTION_HEADERS:
    ws.merge_cells(span)
    cell = ws[span.split(":")[0]]
    cell.value = label
    cell.font  = WHITE
    cell.fill  = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
ws.row_dimensions[3].height = 22

# Column headers (row 4)
HEADERS = [
    "ID","Subsystem","Component","Function",
    "Failure Mode","Failure Effect",
    "Severity (S)","Potential Cause","Occurrence (O)",
    "Prevention Control","Detection Control","Detection (D)",
    "RPN",
    "Recommended Action","Action Owner",
    "Target Date",
    "Rev. Severity","Rev. Occurrence","Rev. Detection","Rev. RPN",
    "Status"
]
COL_WIDTHS = [5,12,18,22,22,28,10,28,10,28,28,10,8,32,12,12,10,10,10,8,12]

for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=4, column=ci, value=h)
    cell.font = WHITE
    cell.fill = SUBHDR_FILL
    cell.alignment = CENTER
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[4].height = 36

# Data rows
for ri, row in df.iterrows():
    excel_row = ri + 5
    fill = ALT_FILL if ri % 2 == 0 else ALT_FILL2
    values = list(row)
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=excel_row, column=ci, value=val)
        cell.fill = fill
        cell.font = NORMAL
        cell.border = BORDER
        cell.alignment = CENTER if ci in [1,7,9,12,13,16,17,18,19,20] else LEFT

    # Color-code RPN column (col 13)
    rpn_cell = ws.cell(row=excel_row, column=13)
    rpn = row["rpn"]
    if rpn >= 150:
        rpn_cell.fill = RED_FILL
        rpn_cell.font = Font(color="FFFFFF", bold=True, name="Arial", size=9)
    elif rpn >= 100:
        rpn_cell.fill = AMBER_FILL
        rpn_cell.font = Font(color="0D1117", bold=True, name="Arial", size=9)
    else:
        rpn_cell.fill = GREEN_FILL
        rpn_cell.font = Font(color="0D1117", bold=True, name="Arial", size=9)

    # Color-code Severity col (7)
    sev = row["severity"]
    sev_cell = ws.cell(row=excel_row, column=7)
    if sev >= 9:
        sev_cell.fill = RED_FILL
        sev_cell.font = Font(color="FFFFFF", bold=True, name="Arial", size=9)
    elif sev >= 7:
        sev_cell.fill = AMBER_FILL
        sev_cell.font = Font(color="0D1117", bold=True, name="Arial", size=9)

    # Revised RPN col (20)
    rev_rpn = row["revised_rpn"]
    rr_cell = ws.cell(row=excel_row, column=20)
    if rev_rpn >= 100:
        rr_cell.fill = AMBER_FILL
        rr_cell.font = Font(color="0D1117", bold=True, name="Arial", size=9)
    elif rev_rpn < 50:
        rr_cell.fill = GREEN_FILL
        rr_cell.font = Font(color="0D1117", bold=True, name="Arial", size=9)

    # Status color (col 21)
    status_cell = ws.cell(row=excel_row, column=21)
    status_colors = {
        "Complete": ("3FB950","0D1117"), "In Progress": ("388BFD","FFFFFF"),
        "Open":     ("F85149","FFFFFF"),
    }
    sc = status_colors.get(row["status"], ("555555","FFFFFF"))
    status_cell.fill = PatternFill("solid", fgColor=sc[0])
    status_cell.font = Font(color=sc[1], bold=True, name="Arial", size=9)

    ws.row_dimensions[excel_row].height = 42

ws.freeze_panes = "E5"

# ── Sheet 2: Top 10 Risks ────────────────────────────────────────────────────
ws2 = wb.create_sheet("Top 10 Risks")
ws2.sheet_properties.tabColor = "F85149"

top10 = df.nlargest(10, "rpn")[["fmea_id","component","failure_mode","severity","occurrence","detection","rpn","recommended_action","status"]].reset_index(drop=True)

ws2["A1"] = "TOP 10 RISK ITEMS BY RPN"
ws2["A1"].font = Font(color="F85149", bold=True, name="Arial", size=13)
ws2["A1"].fill = PatternFill("solid", fgColor="0D1117")
ws2.merge_cells("A1:I1")
ws2["A1"].alignment = CENTER
ws2.row_dimensions[1].height = 28

t10_headers = ["Rank","ID","Component","Failure Mode","Severity","Occurrence","Detection","RPN","Recommended Action"]
t10_widths  = [6,5,18,28,10,10,10,8,40]
for ci, (h, w) in enumerate(zip(t10_headers, t10_widths), 1):
    cell = ws2.cell(row=2, column=ci, value=h)
    cell.font = WHITE
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[2].height = 30

for ri, row in top10.iterrows():
    r = ri + 3
    ws2.cell(row=r, column=1, value=ri+1).font = Font(color="F85149", bold=True, name="Arial", size=10)
    for ci, val in enumerate(row.values, 2):
        cell = ws2.cell(row=r, column=ci, value=val)
        cell.font = NORMAL
        cell.fill = ALT_FILL if ri % 2 == 0 else ALT_FILL2
        cell.border = BORDER
        cell.alignment = CENTER if ci in [1,2,5,6,7,8] else LEFT
    rpn_c = ws2.cell(row=r, column=8)
    rpn_c.fill = RED_FILL
    rpn_c.font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    ws2.row_dimensions[r].height = 36

# ── Sheet 3: RPN Summary ─────────────────────────────────────────────────────
ws3 = wb.create_sheet("RPN Summary")
ws3.sheet_properties.tabColor = "D29922"

comp_summary = df.groupby("component").agg(
    count=("rpn","count"),
    avg_rpn=("rpn","mean"),
    max_rpn=("rpn","max"),
    avg_rev_rpn=("revised_rpn","mean"),
).round(1).reset_index().sort_values("max_rpn", ascending=False)

ws3["A1"] = "RPN SUMMARY BY COMPONENT"
ws3["A1"].font = Font(color="D29922", bold=True, name="Arial", size=13)
ws3["A1"].fill = PatternFill("solid", fgColor="0D1117")
ws3.merge_cells("A1:F1")
ws3["A1"].alignment = CENTER
ws3.row_dimensions[1].height = 28

s3_headers = ["Component","# Failure Modes","Avg RPN (Before)","Max RPN","Avg RPN (After)","Risk Reduction %"]
s3_widths   = [22,16,16,12,16,16]
for ci, (h, w) in enumerate(zip(s3_headers, s3_widths), 1):
    cell = ws3.cell(row=2, column=ci, value=h)
    cell.font = WHITE
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.row_dimensions[2].height = 30

for ri, row in enumerate(comp_summary.itertuples(index=False)):
    row = dict(zip(comp_summary.columns, row))
    r = ri + 3
    pct = round((row["avg_rpn"] - row["avg_rev_rpn"]) / row["avg_rpn"] * 100, 1) if row["avg_rpn"] > 0 else 0
    vals = [row["component"], row["count"], row["avg_rpn"], row["max_rpn"], row["avg_rev_rpn"], f"{pct}%"]
    for ci, val in enumerate(vals, 1):
        cell = ws3.cell(row=r, column=ci, value=val)
        cell.font = NORMAL
        cell.fill = ALT_FILL if (r % 2 == 0) else ALT_FILL2
        cell.border = BORDER
        cell.alignment = LEFT if ci == 1 else CENTER
    ws3.row_dimensions[r].height = 22

# ── Sheet 4: Scoring Guide ───────────────────────────────────────────────────
ws4 = wb.create_sheet("Scoring Guide")
ws4.sheet_properties.tabColor = "3FB950"

def add_guide_section(ws, start_row, title, headers, rows, title_color="58A6FF"):
    ws.merge_cells(f"A{start_row}:C{start_row}")
    ws[f"A{start_row}"] = title
    ws[f"A{start_row}"].font = Font(color=title_color, bold=True, name="Arial", size=11)
    ws[f"A{start_row}"].fill = PatternFill("solid", fgColor="0D1117")
    ws[f"A{start_row}"].alignment = CENTER
    ws.row_dimensions[start_row].height = 22

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row+1, column=ci, value=h)
        cell.font = WHITE
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.row_dimensions[start_row+1].height = 20

    for ri, row in enumerate(rows):
        r = start_row + 2 + ri
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = NORMAL
            cell.fill = ALT_FILL if ri % 2 == 0 else ALT_FILL2
            cell.border = BORDER
            cell.alignment = LEFT if ci == 3 else CENTER
        ws.row_dimensions[r].height = 18

add_guide_section(ws4, 1, "SEVERITY SCALE",
    ["Score","Level","Description"],
    [(10,"Critical","Safety-critical vehicle failure; regulatory violation"),
     (9,"Critical","Mission failure; loss of all propulsion"),
     ("7-8","High","Major performance degradation; significant rework"),
     ("4-6","Moderate","Moderate performance loss; rework required"),
     ("1-3","Low","Minor issue; cosmetic; no functional impact")])

add_guide_section(ws4, 9, "OCCURRENCE SCALE",
    ["Score","Frequency","Description"],
    [(10,"Very High","> 1 in 10 units"),
     ("7-9","High","1 in 20 – 1 in 10"),
     ("4-6","Moderate","1 in 100 – 1 in 20"),
     ("2-3","Low","1 in 1000 – 1 in 100"),
     (1,"Remote","< 1 in 10,000")])

add_guide_section(ws4, 17, "DETECTION SCALE",
    ["Score","Ability","Description"],
    [(10,"None","No current detection control"),
     ("7-9","Weak","Detection only after failure in field"),
     ("4-6","Moderate","Manual inspection with some chance of escape"),
     ("2-3","Strong","Automated in-process or EOL detection"),
     (1,"Very Strong","100% automated detection; mistake-proofed")])

wb.save((DATA_DIR / "propulsion_fmea.xlsx"))
print("✓ propulsion_fmea.xlsx")

# ─── SUMMARY REPORT ──────────────────────────────────────────────────────────

high_rpn      = df[df["rpn"] >= 150].sort_values("rpn", ascending=False)
top10_md      = df.nlargest(10, "rpn")
avg_reduction = ((df["rpn"] - df["revised_rpn"]) / df["rpn"] * 100).mean()
comp_rpn      = df.groupby("component")["rpn"].mean().sort_values(ascending=False)

def divider():
    return "\n---\n"

def meta_table(rows):
    lines = ["| | |", "|---|---|"]
    for k, v in rows:
        lines.append(f"| {k} | {v} |")
    return "\n".join(lines)

# ── Summary report ────────────────────────────────────────────────────────────

comp_table = "\n".join(f"| {c} | {n} |" for c, n in df.groupby("component").size().items())
top10_table = "\n".join(
    f"| {i} | {r['component']} | {r['failure_mode']} | {r['severity']} | {r['occurrence']} | {r['detection']} | {r['rpn']} |"
    for i, (_, r) in enumerate(top10_md.iterrows(), 1)
)
comp_rpn_table = "\n".join(f"| {c} | {a:.1f} |" for c, a in comp_rpn.items())
still_high = df[df["revised_rpn"] > 50].sort_values("revised_rpn", ascending=False)
still_high_lines = "\n".join(
    f"- {r['failure_mode']} ({r['component']}): revised RPN {r['revised_rpn']}"
    for _, r in still_high.head(5).iterrows()
)

summary_md = f"""# FMEA Summary Report
## Drone Propulsion System — Rev A

{meta_table([
    ("Document",     "Propulsion FMEA Rev A"),
    ("Date",         today.isoformat()),
    ("Prepared by",  "Quality Engineering"),
    ("Scope",        "Autonomous Delivery Drone — Propulsion Subsystem"),
    ("RPN Formula",  "Severity × Occurrence × Detection (1–10 scale)"),
])}
{divider()}
## 1. Scope

{len(df)} failure modes analysed across {df['component'].nunique()} propulsion components.

| Component | Failure Modes |
|-----------|--------------|
{comp_table}
{divider()}
## 2. Assumptions

- Analysis covers both design-intent and manufacturing-process failure modes.
- All scores reflect current-state controls — no proposed mitigations applied.
- RPN threshold for high-risk classification: 150.
- Detection controls reference inspection and EOL test capabilities as of {today.isoformat()}.
{divider()}
## 3. Scoring Methodology

RPN = Severity × Occurrence × Detection

| Band | RPN Range | Action Required |
|------|-----------|----------------|
| High | ≥ 150 | Immediate corrective action |
| Moderate | 80–149 | Action plan required |
| Low | < 80 | Monitor; no immediate action |
{divider()}
## 4. Risk Summary

{meta_table([
    ("Total failure modes",        len(df)),
    ("High-risk items (RPN ≥ 150)",len(high_rpn)),
    ("Average RPN — before",       f"{df['rpn'].mean():.1f}"),
    ("Average RPN — after",        f"{df['revised_rpn'].mean():.1f}"),
    ("Average RPN reduction",      f"{avg_reduction:.1f}%"),
    ("Highest RPN",                f"{df['rpn'].max()} — {df.loc[df['rpn'].idxmax(),'failure_mode']}"),
])}
{divider()}
## 5. Top 10 Risks by RPN

| Rank | Component | Failure Mode | S | O | D | RPN |
|------|-----------|-------------|---|---|---|-----|
{top10_table}
{divider()}
## 6. Risk by Component

| Component | Avg RPN |
|-----------|---------|
{comp_rpn_table}
{divider()}
## 7. Recommended Mitigations

Six themes emerged across corrective actions:

- Barcode scan verification at assembly — eliminates wrong-part installation at motor, propeller, and ESC stations
- AQL tightening and CMM SPC — closes dimensional gaps on Motor Mount Bracket hole positions
- Connector upgrades — polarized and locking connectors eliminate reverse polarity and signal-loss risk
- Digital torque logging — real-time pass/fail capture replaces manual torque audit
- Thermal management — TIM shim templates, ESC margin increase, thermal cycle qualification screen
- PMI at receiving — closes material verification gap on high-severity, low-detection bracket lots
{divider()}
## 8. Revised RPN Results

Average RPN reduced from {df['rpn'].mean():.1f} to {df['revised_rpn'].mean():.1f} — a {avg_reduction:.1f}% reduction across the subsystem.

Items still above RPN 50 after mitigation:

{still_high_lines}
{divider()}
## 9. Lessons Learned

- Connector design changes (polarization, locking retention) deliver high RPN reduction at low implementation cost
- Wire routing and chafing protection gaps are underspecified in current SOPs — high-severity, high-occurrence risk
- PMI gaps at receiving represent a high-severity, low-detection condition that cannot be deferred
- Firmware version control requires formal configuration locking in the build traveler, not just release notes review
{divider()}
## 10. Connection to Manufacturing Quality

| Downstream System | How This FMEA Feeds It |
|------------------|------------------------|
| Inspection Plans | High S×O items trigger increased AQL sampling rates |
| NCR Defect Library | Failure modes become standardised defect type categories |
| Supplier Quality | Supplier-caused failures generate SCARs |
| EOL Test Cases | High-detection-score items become new automated EOL checks |
| GD&T Inspection | Bracket dimensional failures define CMM feature requirements |

---
*Generated {today.isoformat()} — Drone Manufacturing Quality Engineering*
"""

with open(REPORTS_DIR / "fmea_summary_report.md", "w") as f:
    f.write(summary_md)
print("✓ fmea_summary_report.md")

# ── Top 10 risks report ───────────────────────────────────────────────────────

top10_lines = ["# Top 10 FMEA Risk Items", "## Drone Propulsion System — Ranked by RPN", "",
               meta_table([("Document","Propulsion FMEA Rev A"),("Date",today.isoformat()),
                            ("Risk threshold","RPN ≥ 150 = High"),("Total failure modes",len(df))]),
               divider()]

for i, (_, row) in enumerate(top10_md.iterrows(), 1):
    reduction_pts = row["rpn"] - row["revised_rpn"]
    reduction_pct = reduction_pts / row["rpn"] * 100
    risk_band = "HIGH" if row["rpn"] >= 150 else "MODERATE"
    top10_lines += [
        f"## Rank {i} — {row['failure_mode']}",
        "",
        meta_table([
            ("Component",        row["component"]),
            ("Risk Band",        risk_band),
            ("RPN",              f"{row['rpn']} (S={row['severity']} × O={row['occurrence']} × D={row['detection']})"),
            ("Failure Effect",   row["failure_effect"]),
            ("Potential Cause",  row["potential_cause"]),
            ("Recommended Action",row["recommended_action"]),
            ("Owner",            row["action_owner"]),
            ("Due Date",         row["target_completion_date"]),
            ("Revised RPN",      f"{row['revised_rpn']} — reduction of {reduction_pts} points ({reduction_pct:.0f}%)"),
            ("Status",           row["status"]),
        ]),
        divider(),
    ]

with open(REPORTS_DIR / "top_10_risks.md", "w") as f:
    f.write("\n".join(top10_lines))
print("✓ top_10_risks.md")

# ── Recommended actions report ────────────────────────────────────────────────

actions          = df[df["recommended_action"].notna()].copy()
actions_by_owner = actions.groupby("action_owner")
complete         = len(actions[actions["status"] == "Complete"])
in_progress      = len(actions[actions["status"] == "In Progress"])
open_count       = len(actions[actions["status"] == "Open"])

action_lines = [
    "# FMEA Recommended Actions",
    "## Drone Propulsion System — Rev A",
    "",
    meta_table([
        ("Document",    "Propulsion FMEA Rev A"),
        ("Date",        today.isoformat()),
        ("Total Actions",len(actions)),
        ("Complete",    complete),
        ("In Progress", in_progress),
        ("Open",        open_count),
    ]),
    divider(),
]

for owner, group in actions_by_owner:
    action_lines += [f"## {owner}", ""]
    action_lines += [
        "| ID | Component | Failure Mode | Action | Due Date | Status |",
        "|----|-----------|-------------|--------|----------|--------|",
    ]
    for _, row in group.iterrows():
        action_lines.append(
            f"| {row['fmea_id']} | {row['component']} | {row['failure_mode']} "
            f"| {row['recommended_action']} | {row['target_completion_date']} | {row['status']} |"
        )
    action_lines.append("")

action_lines += [divider(), f"*Generated {today.isoformat()} — Drone Manufacturing Quality Engineering*"]

with open(REPORTS_DIR / "recommended_actions.md", "w") as f:
    f.write("\n".join(action_lines))
print("✓ recommended_actions.md")

print(f"\nProject 1 complete — {len(df)} failure modes, {len(high_rpn)} high-risk items")
print(f"Avg RPN before: {df['rpn'].mean():.1f} → after: {df['revised_rpn'].mean():.1f} ({avg_reduction:.1f}% reduction)")